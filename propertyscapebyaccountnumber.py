import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# List of property IDs to process
property_ids = [  FILL THIS IN    ]

# Prepare to accumulate data for all properties
all_data = []

# Iterate over each property ID
for property_id in property_ids:
    url = f'https://www.dallascad.org/AcctDetailCom.aspx?ID={property_id}#Legal'

    # Perform a GET request to fetch the HTML content
    page = requests.get(url)

    # Check if the request was successful
    if page.status_code != 200:
        print(f"Failed to retrieve webpage for {property_id}: {page.status_code}")
        continue  # Skip to the next property ID

    # Parse the HTML content
    soup = BeautifulSoup(page.content, "html.parser")

    # Extract the address
    address = soup.find('span', id='PropAddr1_lblPropAddr')
    address_found = address.text.strip() if address else "Address not found"

    # Extract the deed transfer date
    deed_transfer_date = soup.find('span', id='LegalDesc1_lblSaleDate')
    deed_transfer_date_found = deed_transfer_date.text.strip() if deed_transfer_date else "Deed transfer date not found"

    # Extract owner information
    owner_section = soup.find('span', id='lblOwner')
    if owner_section:
        # Initialize a list to collect owner lines
        owner_info_list = []
        for sibling in owner_section.next_siblings:
            if sibling.name == 'br':  # Skip <br> tags
                continue
            if sibling.string:  # Add text content of sibling nodes
                line = sibling.strip().replace('\u00a0', ' ')  # Replace non-breaking spaces with regular spaces
                owner_info_list.append(line)
        owner_found = '\n'.join(owner_info_list)
    else:
        owner_found = "Owner not found"

    # Extract total area
    total_area_display = "Total area not found"
    improvements_table = soup.find('span', string=lambda t: t and 'Improvements (Current 2025)' in t)
    if improvements_table:
        improvements_table = improvements_table.find_next('table')
        for row in improvements_table.find_all('tr'):
            if 'Total Area:' in row.text:
                total_area_display = row.find_all('td')[1].text.strip()  # Get the relevant cell
                break

    # Extract zoning information
    zoning_found = "Zoning info not found"
    land_table = soup.find('table', id='Land1_dgLand')
    if land_table:
        zoning_row = land_table.find_all('tr')
        if len(zoning_row) > 1:
            zoning_cell = zoning_row[1].find_all('td')[2]
            zoning_found = zoning_cell.text.strip() if zoning_cell else zoning_found

    # Extract land area
    land_area_span = soup.find('span', id='Land1_dgLand__ctl2_Label1')
    land_area_found = land_area_span.text.strip() if land_area_span else "Land area not found"

    # Prepare data for this property
    data = {
        'Commercial Account #': property_id,
        'Address': address_found,
        'Deed Transfer Date': deed_transfer_date_found,
        'Owner': owner_found,
        'Total Area': total_area_display,
        'Zoning': zoning_found,
        'Land Area': land_area_found
    }

    # Append this property's data to the all_data list
    all_data.append(data)

# Define the Excel file name
excel_file = "property_datanew.xlsx"

# Create a DataFrame from all collected data
df_all = pd.DataFrame(all_data)

# Check if the file exists
if os.path.exists(excel_file):
    # Load the existing Excel file and append the new data to it
    df_existing = pd.read_excel(excel_file)
    df_combined = pd.concat([df_existing, df_all], ignore_index=True)
    df_combined.to_excel(excel_file, index=False)
else:
    # Create a new Excel file with the collected data
    df_all.to_excel(excel_file, index=False)

# Use openpyxl to open the new Excel file and set text wrapping and column widths
wb = load_workbook(excel_file)
ws = wb.active

# Set wrap_text for owner and address columns
for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(data)):  # Adjust max_col based on your DataFrame
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

# Adjust the width of each column
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter  # Get the column letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    # Set the width to max_length + 2 for some padding
    ws.column_dimensions[column_letter].width = max_length + 2

# Save the workbook
wb.save(excel_file)

print("Data for all properties has been written to Excel successfully.")
